#!/usr/bin/env python3
"""
Genera los CSV de hashes y el JSON de bloqueo de IPs en una sola carpeta por comunicado.

Reglas de negocio de HASHES:
- Busca archivos .xlsx cuyo nombre contenga COMUNICADO-<numero>.
- Lee la hoja "HASH" con encabezados en la fila 4.
- Excluye filas donde POSITIVOS sea 0.
- Genera SIEMPRE 6 CSV por comunicado:
    * SHA1_windows_<numero>.csv
    * SHA1_linux_<numero>.csv
    * SHA1_macos_<numero>.csv
    * SHA256_windows_<numero>.csv
    * SHA256_linux_<numero>.csv
    * SHA256_macos_<numero>.csv
- TODOS los CSV tienen la misma cantidad de hashes válidos.
- No clasifica por SO: duplica el mismo conjunto para windows, linux y macos.
- Description = nombre del Excel sin extensión.
- Source = user.
- Salida = una sola carpeta Salidas_comunicado_<numero> al lado del Excel.

Reglas de negocio de IPs / SentinelOne Firewall Control:
- Lee la hoja "IP", aunque en el Excel venga como "IP " con espacios.
- Usa encabezados en la fila 4.
- Toma la columna "DIRECCION IP".
- Normaliza IPs defangadas, por ejemplo:
    172[.]86[.]113[.]102 -> 172.86.113.102
- Elimina duplicados conservando el orden.
- Genera un JSON con UNA regla por comunicado.
- La regla queda con:
    * action = Block
    * direction = any
    * os_types = linux, osx, windows
    * remote_host_type = addresses
    * remote_hosts = una entrada por IP, siguiendo el formato exportado por SentinelOne
    * status = Disabled por defecto, porque así se viene manejando en consola
- Salida = la misma carpeta Salidas_comunicado_<numero> usada para los CSV de hashes.

Comportamiento:
- Si se pasa un .xlsx por argumento, procesa ese archivo directamente.
- Si no se pasa argumento, escanea la carpeta donde está el .py.
- Si hay un solo comunicado, lo procesa sin preguntar.
- Si hay varios, permite procesar todos o elegir uno por número.
- Por defecto genera hashes + JSON de IPs.
- Opcionalmente se puede usar:
    --hashes-only  para generar solo CSV de hashes.
    --ips-only     para generar solo JSON de IPs.
    --enabled      para generar la regla de IP con status Enabled.
- Mantiene la terminal abierta al final.

Uso:
    python generar_hashes_e_ips_sentinelone.py
    python generar_hashes_e_ips_sentinelone.py "COMUNICADO-1523-ID1095-NOTICIAS_DIARIAS-23042026.xlsx"
    python generar_hashes_e_ips_sentinelone.py archivo.xlsx --enabled
    python generar_hashes_e_ips_sentinelone.py archivo.xlsx --hashes-only
    python generar_hashes_e_ips_sentinelone.py archivo.xlsx --ips-only

Dependencia:
- openpyxl
"""

from __future__ import annotations

import argparse
import ipaddress
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

HASH_SHEET_NAME = "HASH"
IP_SHEET_NAME = "IP"
HEADER_ROW_NUMBER = 4
SOURCE_VALUE = "user"
MAX_IPS_POR_REGLA = 49
COMUNICADO_REGEX = re.compile(r"COMUNICADO[-_ ]*(\d+)", re.IGNORECASE)
IPV4_REGEX = re.compile(r"(?<![\d.])(?:\d{1,3}\.){3}\d{1,3}(?![\d.])")


class ComunicadoError(Exception):
    """Error controlado de procesamiento."""


def normalizar_texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).replace("\xa0", " ").strip()


def normalizar_header(valor: object) -> str:
    texto = normalizar_texto(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    texto = texto.replace("_", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.upper()


def extraer_numero_comunicado(nombre_archivo: str) -> str:
    match = COMUNICADO_REGEX.search(nombre_archivo)
    if not match:
        raise ComunicadoError(
            f"No se pudo extraer el número de comunicado desde: {nombre_archivo}"
        )
    return match.group(1)


def es_comunicado_excel(ruta: Path) -> bool:
    return (
        ruta.is_file()
        and ruta.suffix.lower() == ".xlsx"
        and not ruta.name.startswith("~$")
        and COMUNICADO_REGEX.search(ruta.name) is not None
    )


def convertir_numero(valor: object) -> float | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None

    texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def buscar_hoja_por_nombre(wb, nombre_objetivo: str) -> str:
    for nombre in wb.sheetnames:
        if nombre.strip().upper() == nombre_objetivo.upper():
            return nombre
    raise ComunicadoError(
        f"No existe la hoja '{nombre_objetivo}' en el Excel. Hojas disponibles: {wb.sheetnames}"
    )


def construir_carpeta_salida_unica(ruta_excel: Path) -> Path:
    numero_comunicado = extraer_numero_comunicado(ruta_excel.name)
    return ruta_excel.parent / f"Salidas_comunicado_{numero_comunicado}"


# =========================
# HASHES
# =========================


def leer_hoja_hash(ruta_excel: Path) -> tuple[list[dict[str, str]], str, str]:
    descripcion = ruta_excel.stem
    numero_comunicado = extraer_numero_comunicado(ruta_excel.name)

    try:
        wb = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
    except Exception as exc:
        raise ComunicadoError(f"No se pudo abrir el Excel: {exc}") from exc

    try:
        hoja_hash = buscar_hoja_por_nombre(wb, HASH_SHEET_NAME)
        ws = wb[hoja_hash]

        headers: list[str] = []
        for column_index in range(1, ws.max_column + 1):
            valor = ws.cell(row=HEADER_ROW_NUMBER, column=column_index).value
            headers.append(normalizar_header(valor))

        indice_por_columna = {header: idx for idx, header in enumerate(headers) if header}
        columnas_requeridas = {"SHA1", "SHA256", "POSITIVOS"}
        faltantes = sorted(col for col in columnas_requeridas if col not in indice_por_columna)
        if faltantes:
            raise ComunicadoError(
                f"Faltan columnas requeridas en la hoja '{hoja_hash}': {faltantes}"
            )

        filas_validas: list[dict[str, str]] = []
        faltan_sha1 = 0
        faltan_sha256 = 0

        for fila in ws.iter_rows(min_row=HEADER_ROW_NUMBER + 1, values_only=True):
            if fila is None or all(celda is None or str(celda).strip() == "" for celda in fila):
                continue

            positivos_valor = fila[indice_por_columna["POSITIVOS"]]
            positivos = convertir_numero(positivos_valor)
            if positivos == 0:
                continue

            sha1 = normalizar_texto(fila[indice_por_columna["SHA1"]])
            sha256 = normalizar_texto(fila[indice_por_columna["SHA256"]])

            if not sha1:
                faltan_sha1 += 1
            if not sha256:
                faltan_sha256 += 1

            if not sha1 or not sha256:
                continue

            filas_validas.append(
                {
                    "Description": descripcion,
                    "SHA1": sha1,
                    "SHA256": sha256,
                    "Source": SOURCE_VALUE,
                }
            )

        if faltan_sha1 or faltan_sha256:
            raise ComunicadoError(
                "Hay filas válidas sin hash completo. "
                f"Faltan SHA1: {faltan_sha1}, faltan SHA256: {faltan_sha256}. "
                "No se generaron los CSV para evitar cantidades inconsistentes."
            )

        return filas_validas, descripcion, numero_comunicado
    finally:
        wb.close()


def construir_registros_csv(
    filas_validas: Iterable[dict[str, str]], algoritmo: str, os_name: str
) -> list[dict[str, str]]:
    algoritmo = algoritmo.upper()
    if algoritmo not in {"SHA1", "SHA256"}:
        raise ComunicadoError(f"Algoritmo no soportado: {algoritmo}")

    registros: list[dict[str, str]] = []
    for fila in filas_validas:
        registros.append(
            {
                "OS": os_name,
                "Description": fila["Description"],
                "SHA1": fila["SHA1"] if algoritmo == "SHA1" else "",
                "SHA256": fila["SHA256"] if algoritmo == "SHA256" else "",
                "Source": fila["Source"],
            }
        )
    return registros


def _escape_csv(valor: object) -> str:
    return str(valor).replace('"', '""')


def escribir_csv(ruta_salida: Path, registros: Iterable[dict[str, str]]) -> None:
    """
    Copia el formato del adjunto:
    - OS sin comillas.
    - Description, SHA1, SHA256 y Source siempre entre comillas.
    - Vacíos como "".
    """
    with ruta_salida.open("w", newline="", encoding="utf-8-sig") as archivo_csv:
        archivo_csv.write('OS,"Description","SHA1","SHA256","Source"\n')
        for registro in registros:
            os_value = _escape_csv(registro.get("OS", ""))
            description = _escape_csv(registro.get("Description", ""))
            sha1 = _escape_csv(registro.get("SHA1", ""))
            sha256 = _escape_csv(registro.get("SHA256", ""))
            source = _escape_csv(registro.get("Source", ""))
            archivo_csv.write(
                f'{os_value},"{description}","{sha1}","{sha256}","{source}"\n'
            )


def guardar_csvs_hashes(ruta_excel: Path, carpeta_salida: Path | None = None) -> list[Path]:
    filas_validas, descripcion, numero_comunicado = leer_hoja_hash(ruta_excel)

    if carpeta_salida is None:
        carpeta_salida = construir_carpeta_salida_unica(ruta_excel)
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    archivos_generados: list[Path] = []
    combinaciones = [
        ("SHA1", "windows"),
        ("SHA1", "linux"),
        ("SHA1", "macos"),
        ("SHA256", "windows"),
        ("SHA256", "linux"),
        ("SHA256", "macos"),
    ]

    for algoritmo, os_name in combinaciones:
        registros = construir_registros_csv(filas_validas, algoritmo, os_name)
        nombre_archivo = f"{algoritmo}_{os_name}_{numero_comunicado}.csv"
        ruta_salida = carpeta_salida / nombre_archivo
        escribir_csv(ruta_salida, registros)
        archivos_generados.append(ruta_salida)

    print(f"\n[HASHES] Procesado: {ruta_excel.name}")
    print(f"  Descripción: {descripcion}")
    print(f"  Número de comunicado: {numero_comunicado}")
    print(f"  Filas válidas exportadas por CSV: {len(filas_validas)}")
    print(f"  Carpeta de salida: {carpeta_salida}")
    print("  Archivos generados:")
    for archivo in archivos_generados:
        print(f"    - {archivo.name}")

    return archivos_generados


# =========================
# IPS / SENTINELONE
# =========================


def desofuscar_texto_ip(texto: str) -> str:
    """Convierte variantes comunes de IP defangada a formato normal."""
    texto = normalizar_texto(texto)
    reemplazos = {
        "[.]": ".",
        "(.)": ".",
        "{.}": ".",
        "[dot]": ".",
        "(dot)": ".",
        "hxxp://": "http://",
        "hxxps://": "https://",
    }
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
        texto = texto.replace(origen.upper(), destino)
    return texto


def extraer_ips_desde_valor(valor: object) -> list[str]:
    texto = desofuscar_texto_ip(normalizar_texto(valor))
    if not texto:
        return []

    ips: list[str] = []
    for candidato in IPV4_REGEX.findall(texto):
        try:
            ip = str(ipaddress.ip_address(candidato))
        except ValueError:
            continue
        ips.append(ip)
    return ips


def deduplicar_preservando_orden(valores: Iterable[str]) -> list[str]:
    vistos: set[str] = set()
    resultado: list[str] = []
    for valor in valores:
        if valor not in vistos:
            vistos.add(valor)
            resultado.append(valor)
    return resultado


def leer_ips_comunicado(ruta_excel: Path) -> tuple[list[str], str, str, list[str]]:
    nombre_regla = ruta_excel.stem
    numero_comunicado = extraer_numero_comunicado(ruta_excel.name)

    try:
        wb = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
    except Exception as exc:
        raise ComunicadoError(f"No se pudo abrir el Excel: {exc}") from exc

    try:
        hoja_ip = buscar_hoja_por_nombre(wb, IP_SHEET_NAME)
        ws = wb[hoja_ip]

        headers: list[str] = []
        for column_index in range(1, ws.max_column + 1):
            valor = ws.cell(row=HEADER_ROW_NUMBER, column=column_index).value
            headers.append(normalizar_header(valor))

        indice_por_columna = {header: idx for idx, header in enumerate(headers) if header}

        columna_ip = "DIRECCION IP"
        if columna_ip not in indice_por_columna:
            raise ComunicadoError(
                f"Falta la columna requerida '{columna_ip}' en la hoja '{hoja_ip}'. "
                f"Columnas detectadas: {list(indice_por_columna.keys())}"
            )

        indice_ip = indice_por_columna[columna_ip]

        ips_encontradas: list[str] = []
        filas_sin_ip_valida: list[str] = []

        for numero_fila, fila in enumerate(
            ws.iter_rows(min_row=HEADER_ROW_NUMBER + 1, values_only=True),
            start=HEADER_ROW_NUMBER + 1,
        ):
            if fila is None or all(celda is None or str(celda).strip() == "" for celda in fila):
                continue

            valor_ip = fila[indice_ip] if indice_ip < len(fila) else None
            ips_fila = extraer_ips_desde_valor(valor_ip)

            if ips_fila:
                ips_encontradas.extend(ips_fila)
            else:
                texto_original = normalizar_texto(valor_ip)
                if texto_original:
                    filas_sin_ip_valida.append(f"Fila {numero_fila}: {texto_original}")

        ips_unicas = deduplicar_preservando_orden(ips_encontradas)

        if not ips_unicas:
            raise ComunicadoError(f"No se encontraron IPs válidas en la hoja '{hoja_ip}'.")

        return ips_unicas, nombre_regla, numero_comunicado, filas_sin_ip_valida
    finally:
        wb.close()


def construir_regla_sentinelone(nombre_regla: str, ips: list[str], status: str) -> dict:
    if status not in {"Disabled", "Enabled"}:
        raise ComunicadoError(f"Estado no soportado para la regla: {status}")

    return {
        "action": "Block",
        "application": [],
        "application_type": "any",
        "description": None,
        "direction": "any",
        "local_host": [],
        "local_host_type": "any",
        "local_port": [],
        "local_port_type": "any",
        "location_ids": [],
        "location_type": "all",
        "name": nombre_regla,
        "os_types": ["linux", "osx", "windows"],
        "profile": "any",
        "protocol": None,
        "remote_host": [ips[0]],
        "remote_host_type": "addresses",
        "remote_hosts": [{"type": "addresses", "values": [ip]} for ip in ips],
        "remote_port": [],
        "remote_port_type": "any",
        "rule_type": "custom",
        "scope": "account",
        "service": None,
        "status": status,
        "tag_ids": [],
        "tag_names": [],
    }


def guardar_json_bloqueo_ip(
    ruta_excel: Path, enabled: bool = False, carpeta_salida: Path | None = None
) -> list[Path]:
    ips, nombre_regla, numero_comunicado, filas_sin_ip_valida = leer_ips_comunicado(ruta_excel)

    status = "Enabled" if enabled else "Disabled"

    chunks = [ips[i:i + MAX_IPS_POR_REGLA] for i in range(0, len(ips), MAX_IPS_POR_REGLA)]

    if carpeta_salida is None:
        carpeta_salida = construir_carpeta_salida_unica(ruta_excel)
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    archivos_generados: list[Path] = []

    if len(chunks) == 1:
        regla = construir_regla_sentinelone(nombre_regla=nombre_regla, ips=chunks[0], status=status)
        ruta_salida = carpeta_salida / f"rules_{nombre_regla}.json"
        with ruta_salida.open("w", encoding="utf-8") as archivo_json:
            json.dump([regla], archivo_json, indent=2, ensure_ascii=False)
            archivo_json.write("\n")
        archivos_generados.append(ruta_salida)
    else:
        for idx, chunk in enumerate(chunks, start=1):
            nombre_parte = f"{nombre_regla} ({idx} de {len(chunks)})"
            regla = construir_regla_sentinelone(nombre_regla=nombre_parte, ips=chunk, status=status)
            ruta_salida = carpeta_salida / f"rules_{nombre_regla}_parte_{idx}.json"
            with ruta_salida.open("w", encoding="utf-8") as archivo_json:
                json.dump([regla], archivo_json, indent=2, ensure_ascii=False)
                archivo_json.write("\n")
            archivos_generados.append(ruta_salida)

    print(f"\n[IPS] Procesado: {ruta_excel.name}")
    print(f"  Regla base: {nombre_regla}")
    print(f"  Número de comunicado: {numero_comunicado}")
    print(f"  Estado de la regla: {status}")
    print(f"  IPs únicas totales: {len(ips)}")
    print(f"  Archivos JSON generados: {len(archivos_generados)}")
    print(f"  Carpeta de salida: {carpeta_salida}")
    for archivo in archivos_generados:
        print(f"    - {archivo.name}")

    if filas_sin_ip_valida:
        print("\n  Advertencia: se detectaron valores no vacíos sin IP válida:")
        for item in filas_sin_ip_valida[:20]:
            print(f"    - {item}")
        if len(filas_sin_ip_valida) > 20:
            print(f"    ... y {len(filas_sin_ip_valida) - 20} más.")

    return archivos_generados


# =========================
# FLUJO GENERAL
# =========================


def escanear_comunicados(carpeta_base: Path) -> list[Path]:
    return sorted(
        (ruta for ruta in carpeta_base.iterdir() if es_comunicado_excel(ruta)),
        key=lambda p: p.name.lower(),
    )


def mostrar_lista(comunicados: list[Path]) -> None:
    print("\nComunicados encontrados:")
    for i, ruta in enumerate(comunicados, start=1):
        print(f"  {i}. {ruta.name}")


def preguntar_si_no(mensaje: str) -> bool:
    while True:
        respuesta = input(mensaje).strip().lower()
        if respuesta in {"s", "si", "sí", "y", "yes"}:
            return True
        if respuesta in {"n", "no"}:
            return False
        print("Respuesta no válida. Escribe S o N.")


def elegir_comunicado(comunicados: list[Path]) -> Path:
    while True:
        seleccion = input("Escribe el número del comunicado a procesar: ").strip()
        if not seleccion.isdigit():
            print("Debes escribir un número válido.")
            continue
        indice = int(seleccion)
        if 1 <= indice <= len(comunicados):
            return comunicados[indice - 1]
        print("Número fuera de rango.")


def procesar_comunicado(ruta_excel: Path, generar_hashes: bool, generar_ips: bool, enabled: bool) -> None:
    if not generar_hashes and not generar_ips:
        raise ComunicadoError("No hay ninguna salida seleccionada para generar.")

    print("\n" + "=" * 80)
    print(f"Comunicado: {ruta_excel.name}")
    print("=" * 80)

    carpeta_salida = construir_carpeta_salida_unica(ruta_excel)
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    print(f"Carpeta única de salida: {carpeta_salida}")

    if generar_hashes:
        guardar_csvs_hashes(ruta_excel, carpeta_salida=carpeta_salida)

    if generar_ips:
        guardar_json_bloqueo_ip(ruta_excel, enabled=enabled, carpeta_salida=carpeta_salida)


def procesar_varios(comunicados: list[Path], generar_hashes: bool, generar_ips: bool, enabled: bool) -> int:
    errores = 0
    for comunicado in comunicados:
        try:
            procesar_comunicado(
                comunicado,
                generar_hashes=generar_hashes,
                generar_ips=generar_ips,
                enabled=enabled,
            )
        except Exception as exc:
            errores += 1
            print(f"\nError procesando {comunicado.name}: {exc}")
    return errores


def procesar_argumento(path_str: str, generar_hashes: bool, generar_ips: bool, enabled: bool) -> int:
    ruta = Path(path_str).expanduser().resolve()
    if not ruta.exists():
        print(f"No existe el archivo: {ruta}")
        return 1
    if ruta.suffix.lower() != ".xlsx":
        print(f"El archivo no es .xlsx: {ruta.name}")
        return 1
    try:
        procesar_comunicado(
            ruta,
            generar_hashes=generar_hashes,
            generar_ips=generar_ips,
            enabled=enabled,
        )
        return 0
    except Exception as exc:
        print(f"Error: {exc}")
        return 1


def parsear_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera CSV de hashes y JSON de bloqueo de IPs para comunicados CSIRT/SentinelOne."
    )
    parser.add_argument(
        "excel",
        nargs="?",
        help="Ruta opcional del archivo .xlsx del comunicado.",
    )
    parser.add_argument(
        "--enabled",
        action="store_true",
        help="Genera la regla de bloqueo de IP con status Enabled. Por defecto se genera Disabled.",
    )
    grupo = parser.add_mutually_exclusive_group()
    grupo.add_argument(
        "--hashes-only",
        action="store_true",
        help="Genera solo los CSV de hashes.",
    )
    grupo.add_argument(
        "--ips-only",
        action="store_true",
        help="Genera solo el JSON de bloqueo de IPs.",
    )
    return parser.parse_args()


def main() -> int:
    args = parsear_argumentos()

    generar_hashes = not args.ips_only
    generar_ips = not args.hashes_only

    if args.excel:
        return procesar_argumento(
            args.excel,
            generar_hashes=generar_hashes,
            generar_ips=generar_ips,
            enabled=args.enabled,
        )

    carpeta_base = Path(__file__).resolve().parent
    print(f"Carpeta escaneada: {carpeta_base}")

    comunicados = escanear_comunicados(carpeta_base)
    if not comunicados:
        print("No se encontraron archivos .xlsx de comunicado en la carpeta del script.")
        return 1

    mostrar_lista(comunicados)

    if len(comunicados) == 1:
        print("\nSolo hay un comunicado. Se procesará automáticamente.")
        try:
            procesar_comunicado(
                comunicados[0],
                generar_hashes=generar_hashes,
                generar_ips=generar_ips,
                enabled=args.enabled,
            )
            return 0
        except Exception as exc:
            print(f"Error: {exc}")
            return 1

    if preguntar_si_no("\n¿Deseas generar las salidas para todos los comunicados encontrados? (S/N): "):
        errores = procesar_varios(
            comunicados,
            generar_hashes=generar_hashes,
            generar_ips=generar_ips,
            enabled=args.enabled,
        )
        return 1 if errores else 0

    seleccionado = elegir_comunicado(comunicados)
    try:
        procesar_comunicado(
            seleccionado,
            generar_hashes=generar_hashes,
            generar_ips=generar_ips,
            enabled=args.enabled,
        )
        return 0
    except Exception as exc:
        print(f"Error: {exc}")
        return 1


if __name__ == "__main__":
    try:
        codigo = main()
    except KeyboardInterrupt:
        print("\nOperación cancelada por el usuario.")
        codigo = 1
    except Exception as exc:
        print(f"\nError no controlado: {exc}")
        codigo = 1

    try:
        input("\nPresiona Enter para salir...")
    except EOFError:
        pass
    sys.exit(codigo)
